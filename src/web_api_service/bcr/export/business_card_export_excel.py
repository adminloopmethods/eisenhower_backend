import csv

from django.conf import settings
from django.http import HttpResponse
from openpyxl import Workbook

from cards.models import BusinessCardAddress
from cards.models import BusinessCardDates
from cards.models import BusinessCardEmail
from cards.models import BusinessCardFax
from cards.models import BusinessCardJob
from cards.models import BusinessCardMobile
from cards.models import BusinessCardReaderManager
from cards.models import BusinessCardSocialNetwork
from cards.models import BusinessCardWeb
# api error response parser
from core.api_response_parser import bad_request_response
from core.api_response_parser import final_response
from core.api_response_parser import not_acceptable_response
from web_api_service.bcr.business_card_reader.business_card_serializers import (
    BusinessCardJobSerializer,
)
# all excel header and heading
from web_api_service.bcr.export.business_card_excel_header_titles import (
    BUSINESS_CARD_FIELDS_LIST,
    BUSINESS_CARD_SHEET_TITLE_LIST,
    BUSINESS_CARD_FIELDS,
    BUSINESS_CARD_SHEET_TITLE

)
from web_api_service.helpers.all_config_func import (
    get_auth_instance_from_user_id,
)


# from xlwt import Workbook
# import business card models


class ExportBusinessCards:
    """
    ExportTasksService

    ##
    https://djangotricks.blogspot.com/2019/02/how-to-export-data-to-xlsx-files.html
    ##
    https://blog.devgenius.io/building-your-own-export
    -and-import-data-into-excel-using-django-celery-pandas-%EF%B8%8F-with-784fd688e328
    """

    def __init__(self, **kwargs):
        self.project_base_url = settings.PROJECT_BASE_URL
        self.auth_instance = None
        self.auth_user_instance = kwargs.get("auth_instance", None)
        self.business_card_id = kwargs.get("business_card_id", None)
        self.business_card_qs = BusinessCardReaderManager.objects.filter(
            retrive_status__in=[
                'retrive',
                'failed',
                'discard'
            ]
        )
        self.final_response = final_response()
        self.not_acceptable_response = not_acceptable_response()
        self.bad_request_response = bad_request_response()
        self.business_card_fields_list = BUSINESS_CARD_FIELDS
        self.business_card_sheet_title_list = BUSINESS_CARD_SHEET_TITLE

        self.business_card_fields_details = BUSINESS_CARD_FIELDS_LIST
        self.business_card_sheet_title_details = BUSINESS_CARD_SHEET_TITLE_LIST

        self.file_name = "business_card_list"
        # self.file_content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        # self.file_content_type = 'application/ms-excel'
        self.file_content_type = "application/vnd.ms-excel"
        self.file_content_type = "text/csv"
        self.titles = None
        self.fields = None
        self.export_default_none = 'None'
        self.business_card_list_or_detail = None
        self.columns_list = []
        self.rows_values_list = []
        self.work_book_instance = None
        self.row = []
        self.sheet_type_content = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    def add_company_job_details(self, obj):
        """this add_company_job_details method used to add
        comapany and department and position details using excel
        obj: business_card_obj for using the instance
        """
        try:
            business_card_job_qs = BusinessCardJob.objects.filter(
                business_card_id=obj.id
            )
        except ValueError:
            business_card_job_qs = None
        try:
            job_serializer_data = BusinessCardJobSerializer(
                business_card_job_qs,
                many=True).data
            for _job_data in job_serializer_data:
                try:
                    if _job_data["type"] == "company":
                        if _job_data["job_value"]:
                            self.columns_list.append("company")
                            self.row.append(_job_data["job_value"])
                except Exception as e:
                    print(e)
                try:
                    if _job_data["type"] == "department":
                        if _job_data["job_value"]:
                            self.columns_list.append("department")
                            self.row.append(_job_data["job_value"])
                except Exception as e:
                    print('departmentException')
                    print(e)
                try:
                    if _job_data["type"] == "position":
                        if _job_data["job_value"]:
                            self.columns_list.append("position")
                            self.row.append(_job_data["job_value"])
                except Exception as e:
                    print(e)
                    print('PositionException')
        except Exception as e:
            print("companyJobDetailsExpErr")
            print(e)

        return True

    def add_mobile_details(self, obj):
        """this add_emails_details method used to add emails details using excel
        obj: business_card_obj for using the instance
        """
        try:
            business_card_mobile_qs = BusinessCardMobile.objects.filter(
                business_card_id=obj.id
            )
        except ValueError:
            business_card_mobile_qs = None

        for _mobile_data in business_card_mobile_qs:
            try:
                if _mobile_data.type == "work":
                    self.columns_list.append("mobile(work)")
                    self.row.append(_mobile_data.mobile_value)
            except Exception as e:
                print(e)
            try:
                if _mobile_data.type == "home":
                    self.columns_list.append("mobile(home)")
                    self.row.append(_mobile_data.mobile_value)
            except Exception as e:
                print(e)
            try:
                if _mobile_data.type == "other":
                    self.columns_list.append("mobile(other)")
                    self.row.append(_mobile_data.mobile_value)
            except Exception as e:
                print(e)
            try:
                if _mobile_data.type == "main":
                    self.columns_list.append("mobile(main)")
                    self.row.append(_mobile_data.mobile_value)
            except Exception as e:
                print(e)

        return True

    def add_emails_details(self, obj):
        """this add_emails_details method used to add emails details using excel
        obj: business_card_obj for using the instance
        """
        try:
            business_card_emails_qs = BusinessCardEmail.objects.filter(
                business_card_id=obj.id)
        except ValueError:
            business_card_emails_qs = None

        for _emails_data in business_card_emails_qs:
            try:
                if _emails_data.type == 'work':
                    self.columns_list.append("email(work)")
                    self.row.append(_emails_data.email_value)
            except Exception as e:
                print(e)
            try:
                if _emails_data.type == "home":
                    self.columns_list.append("email(home)")
                    self.row.append(_emails_data.email_value)
            except Exception as e:
                print(e)
            try:
                if _emails_data.type == "other":
                    self.columns_list.append("email(other)")
                    self.row.append(_emails_data.email_value)
            except Exception as e:
                print(e)
            try:
                if _emails_data.type == "main":
                    self.columns_list.append("email(main)")
                    self.row.append(_emails_data.email_value)
            except Exception as e:
                print(e)

        return True

    def add_address_details(self, obj):
        """this add_emails_details method used to add emails details using excel
        obj: business_card_obj for using the instance
        """
        try:
            business_card_address_qs = BusinessCardAddress.objects.filter(
                business_card_id=obj.id)
        except ValueError:
            business_card_address_qs = None

        for _address_data in business_card_address_qs:
            try:
                if _address_data.type == "company":
                    self.columns_list.append("address_company_street")
                    self.row.append(_address_data.street)
                    self.columns_list.append("address_company_city")
                    self.row.append(_address_data.city)
                    self.columns_list.append("address_company_region")
                    self.row.append(_address_data.region)
                    self.columns_list.append("address_company_country")
                    self.row.append(_address_data.country)
                    self.columns_list.append("address_company_zip_code")
                    self.row.append(_address_data.zip_code)
            except Exception as e:
                print(e)
            try:
                if _address_data.type == "home":
                    self.columns_list.append("address_home_street")
                    self.row.append(_address_data.street)
                    self.columns_list.append("address_home_city")
                    self.row.append(_address_data.city)
                    self.columns_list.append("address_home_region")
                    self.row.append(_address_data.region)
                    self.columns_list.append("address_home_country")
                    self.row.append(_address_data.country)
                    self.columns_list.append("address_home_zip_code")
                    self.row.append(_address_data.zip_code)
            except Exception as e:
                print(e)
            try:
                if _address_data.type == "other":
                    self.columns_list.append("address_other_street")
                    self.row.append(_address_data.street)
                    self.columns_list.append("address_other_city")
                    self.row.append(_address_data.city)
                    self.columns_list.append("address_other_region")
                    self.row.append(_address_data.region)
                    self.columns_list.append("address_other_country")
                    self.row.append(_address_data.country)
                    self.columns_list.append("address_other_zip_code")
                    self.row.append(_address_data.zip_code)
            except Exception as e:
                print(e)
            try:
                if _address_data.type == "main":
                    self.columns_list.append("address_main_street")
                    self.row.append(_address_data.street)
                    self.columns_list.append("address_main_city")
                    self.row.append(_address_data.city)
                    self.columns_list.append("address_main_region")
                    self.row.append(_address_data.region)
                    self.columns_list.append("address_main_country")
                    self.row.append(_address_data.country)
                    self.columns_list.append("address_main_zip_code")
                    self.row.append(_address_data.zip_code)
            except Exception as e:
                print(e)

        return True

    def add_dates_details(self, obj):
        """this add_emails_details method used to add emails details using excel
        obj: business_card_obj for using the instance
        """
        try:
            business_card_dates_qs = BusinessCardDates.objects.filter(
                business_card_id=obj.id
            )
        except ValueError:
            business_card_dates_qs = None

        for _dates_data in business_card_dates_qs:
            try:
                if _dates_data.type == "birthday":
                    self.columns_list.append("birthday")
                    self.row.append(_dates_data.date)
            except Exception as e:
                print(e)
            try:
                if _dates_data.type == "anniversary":
                    self.columns_list.append("anniversary")
                    self.row.append(_dates_data.date)
            except Exception as e:
                print(e)
            try:
                if _dates_data.type == "other":
                    self.columns_list.append("other_date")
                    self.row.append(_dates_data.date)

            except Exception as e:
                print(e)
        return True

    def add_webs_details(self, obj):
        """this add_emails_details method used to add emails details using excel
        obj: business_card_obj for using the instance
        """
        try:
            business_card_webs_qs = BusinessCardWeb.objects.filter(
                business_card_id=obj.id
            )
        except ValueError:
            business_card_webs_qs = None

        for _webs_data in business_card_webs_qs:
            try:
                if _webs_data.type == 'work':
                    self.columns_list.append("website(work)")
                    self.row.append(_webs_data.website)
            except Exception as e:
                print(e)
            try:
                if _webs_data.type == 'home':
                    self.columns_list.append("website(home)")
                    self.row.append(_webs_data.website)
            except Exception as e:
                print(e)
            try:
                if _webs_data.type == 'main':
                    self.columns_list.append("website(main)")
                    self.row.append(_webs_data.website)
            except Exception as e:
                print(e)
            try:
                if _webs_data.type == 'other':
                    self.columns_list.append("website(other)")
                    self.row.append(_webs_data.website)
            except Exception as e:
                print(e)
        return True

    def add_fax_details(self, obj):
        """this add_emails_details method used to add emails details using excel
        obj: business_card_obj for using the instance
        """
        try:
            business_card_fax_qs = BusinessCardFax.objects.filter(
                business_card_id=obj.id
            )
        except ValueError:
            business_card_fax_qs = None

        for _webs_data in business_card_fax_qs:
            try:
                if _webs_data.type == 'work':
                    self.columns_list.append("fax(work)")
                    self.row.append(_webs_data.fax_value)
            except Exception as e:
                print(e)
            try:
                if _webs_data.type == 'home':
                    self.columns_list.append("fax(home)")
                    self.row.append(_webs_data.fax_value)
            except Exception as e:
                print(e)
            try:
                if _webs_data.type == 'main':
                    self.columns_list.append("fax(main)")
                    self.row.append(_webs_data.fax_value)
            except Exception as e:
                print(e)
            try:
                if _webs_data.type == 'other':
                    self.columns_list.append("fax(other)")
                    self.row.append(_webs_data.fax_value)
            except Exception as e:
                print(e)
        return True

    def add_social_details(self, obj):
        """this add_emails_details method used to add emails details using excel
        obj: business_card_obj for using the instance
        """
        try:
            business_card_social_qs = BusinessCardSocialNetwork.objects.filter(
                business_card_id=obj.id
            )
        except ValueError:
            business_card_social_qs = None

        for _social_data in business_card_social_qs:
            try:
                if _social_data.type == 'facebook':
                    self.columns_list.append("facebook")
                    self.row.append(_social_data.social_network_value)
            except Exception as e:
                print(e)
            try:
                if _social_data.type == 'linkedin':
                    self.columns_list.append("linkedin")
                    self.row.append(_social_data.social_network_value)
            except Exception as e:
                print(e)
            try:
                if _social_data.type == 'twitter':
                    self.columns_list.append("twitter")
                    self.row.append(_social_data.social_network_value)
            except Exception as e:
                print(e)
            try:
                if _social_data.type == 'googletalk':
                    self.columns_list.append("googletalk")
                    self.row.append(_social_data.social_network_value)
            except Exception as e:
                print(e)

            try:
                if _social_data.type == 'hangout':
                    self.columns_list.append("hangout")
                    self.row.append(_social_data.social_network_value)
            except Exception as e:
                print(e)
            try:
                if _social_data.type == 'skype':
                    self.columns_list.append("skype")
                    self.row.append(_social_data.social_network_value)
            except Exception as e:
                print(e)
            try:
                if _social_data.type == 'yahoo':
                    self.columns_list.append("yahoo")
                    self.row.append(_social_data.social_network_value)
            except Exception as e:
                print(e)
            try:
                if _social_data.type == 'msn':
                    self.columns_list.append("msn")
                    self.row.append(_social_data.social_network_value)
            except Exception as e:
                print(e)
        return True

    def category_name_from_object(self, obj):
        """category_name_from_object
        """
        try:
            obj.category = str(obj.category_group.category_name)
        except Exception as e:
            print(e)
            obj.category = 'None'

    def business_card_image_url(self, obj):
        """business_card_image_url
        """
        try:
            obj.business_image = ''.join(
                [
                    self.project_base_url,
                    '/media/', str(obj.business_image)
                ]
            )
        except Exception as e:
            print('businessImageException')
            print(e)

    def nested_getattr(self, obj, attribute, split_rule="__"):
        """
        This function is responsible for getting the nested record
        from the given obj parameter
        :param obj: whole item without splitting
        :param attribute: field after splitting
        :param split_rule:
        :return:
        """

        if self.business_card_list_or_detail == '__list__':
            self.category_name_from_object(obj)
            self.business_card_image_url(obj)

        if self.business_card_list_or_detail == '__detail__':
            self.add_company_job_details(obj)
            self.add_mobile_details(obj)
            self.add_emails_details(obj)
            self.add_address_details(obj)
            self.add_webs_details(obj)
            self.add_social_details(obj)
            self.add_dates_details(obj)
            self.add_fax_details(obj)

        # all date time split
        try:
            datetime_format = "%d-%m-%Y"
            obj.Created_at = obj.created_at.strftime(datetime_format)
            # obj.created_at = str(obj.created_at).split(" ")[0]
        except Exception as e:
            print('dateSplitFormatException')
            print(e)

        split_attr = attribute.split(split_rule)
        for attr in split_attr:
            if not obj:
                break
            obj = getattr(obj, attr)

        return obj

    def work_book_excel_setting(self):
        """set_work_book_excel_setting
        This set_work_book_excel_setting() methods used to set the work
        book instance for store
        and create excel sheet data for download the payroll user
        also add excel sheet XF Style to set font styles
        """
        # work-book-sheet-instance with
        self.work_book_instance = Workbook()
        # Get active worksheet/tab
        self.work_sheet = self.work_book_instance.active
        self.work_sheet.title = 'business_cards'

    def export_business_cards(self):
        """export_business_cards method used to export
        the all business card detail in Excel format
        Returns: media_excel: Excel sheet file
        """
        try:
            self.auth_user_instance = get_auth_instance_from_user_id(
                self.business_card_id
            )
        except Exception as e:
            self.auth_instance = None
            print("Exception")
            print(e)

        try:
            business_card_queryset = self.business_card_qs.filter(
                user_id=self.business_card_id, is_active=True
            )
            self.business_card_list_or_detail = '__list__'
        except Exception as e:
            print(e)
            business_card_queryset = None

        if not business_card_queryset:
            try:
                business_card_queryset = (
                    self.business_card_qs.filter(
                        id=self.business_card_id, is_active=True
                    )
                    if self.business_card_id
                    else None
                )
                self.business_card_list_or_detail = '__detail__'
            except Exception as e:
                print(e)
                business_card_queryset = None

        if not business_card_queryset:
            business_card_queryset = BusinessCardReaderManager.objects.filter(
                retrive_status='initiate'
            )

        if self.business_card_list_or_detail == '__list__':
            task_fields = self.business_card_fields_list
            self.fields = self.business_card_fields_list
            self.titles = self.business_card_sheet_title_list
            model = business_card_queryset.model
            response = HttpResponse(content_type=self.file_content_type)
            # force download
            response[
                "Content-Disposition"
            ] = "attachment; filename={}.csv".format(
                self.file_name
            )
            # the csv writer
            writer = csv.writer(response)
            if self.fields:
                headers = self.fields
                if self.titles:
                    self.titles = self.titles
                else:
                    self.titles = headers
            else:
                headers = []

                for field in model._meta.fields:
                    headers.append(field.name)
                self.titles = headers

            # Writes the title for the file
            writer.writerow(self.titles)

            # write data rows
            for item in business_card_queryset:
                # a=str(item.start_date)
                writer.writerow(
                    [self.nested_getattr(item, field) for field in headers]
                )

            return response

        if self.business_card_list_or_detail == '__detail__':
            task_fields = self.business_card_fields_details
            self.fields = self.business_card_fields_details
            self.titles = self.business_card_sheet_title_details
            model = business_card_queryset.model

            self.row_query_data = business_card_queryset

            self.work_book_excel_setting()

            # wb = xlwt.Workbook(encoding='utf-8')
            # ws = wb.add_sheet('Tasks')
            # # heading
            # ws.write(0, 0, 'Firstname')
            # ws.write(0, 1, 'Surname')
            # ws.write(0, 2, 'email')
            # ws.write(0, 3, 'company_name')
            # ws.write(0, 4, 'address')
            # # row
            # ws.write(1, 0, 'Gourav')
            # ws.write(1, 1, 'Sharma')
            # ws.write(1, 2, 'pycodertest@gmail.com')
            # ws.write(1, 3, 'loopmethods')
            # ws.write(1, 4, 'Green Park')
            #     "second_name",
            # "last_name",
            # "suffix",
            # "company_name",
            # "telephone_office",
            # "telephone_mobile",
            # "fax",
            # "telephone_home",
            # "e_mail",
            # "country",
            # "city",
            # "street",
            # "postal_code",
            # "web",
            # "facebook",
            # "category_group",

            # Sheet header, first row
            for _obj in business_card_queryset:
                if _obj.business_image:
                    self.columns_list.append("business_image")
                    self.row.append(
                        ''.join(
                            [
                                self.project_base_url,
                                '/media/', str(_obj.business_image)
                            ]
                        )
                    )
                    # self.rows_values_list.append(str(_obj.business_image))
                try:
                    if _obj.first_name:
                        self.columns_list.append("first_name")
                        self.row.append(str(_obj.first_name))
                except Exception as e:
                    print('firstNameExp')
                    print(e)

                try:
                    if _obj.second_name:
                        self.columns_list.append("second_name")
                        self.row.append(str(_obj.second_name))
                except Exception as e:
                    print('firstNameExp')
                    print(e)

                try:
                    if _obj.last_name:
                        self.columns_list.append("last_name")
                        self.row.append(str(_obj.last_name))
                except Exception as e:
                    print('firstNameExp')
                    print(e)

                try:
                    if _obj.suffix:
                        self.columns_list.append("suffix")
                        self.row.append(str(_obj.suffix))
                except Exception as e:
                    print('firstNameExp')
                    print(e)

                try:
                    if _obj.nick_name:
                        self.columns_list.append("nick_name")
                        self.row.append(str(_obj.nick_name))
                except Exception as e:
                    print('firstNameExp')
                    print(e)

                # try:
                #     if _obj.company_name:
                #         self.columns_list.append("company_name")
                #         self.row.append(str(_obj.company_name))
                # except Exception as e:
                #     print('firstNameExp')
                #     print(e)
                #
                # try:
                #     if _obj.telephone_office:
                #         self.columns_list.append("telephone_office")
                #         self.row.append(str(_obj.telephone_office))
                # except Exception as e:
                #     print('firstNameExp')
                #     print(e)
                #
                # try:
                #     if _obj.telephone_mobile:
                #         self.columns_list.append("telephone_mobile")
                #         self.row.append(str(_obj.telephone_mobile))
                # except Exception as e:
                #     print('firstNameExp')
                #     print(e)

                try:
                    if _obj.category_group:
                        self.columns_list.append("category_group")
                        self.row.append(str(_obj.category_group.category_name))
                except Exception as e:
                    print('firstNameExp')
                    print(e)

                self.add_company_job_details(_obj)
                self.add_mobile_details(_obj)
                self.add_emails_details(_obj)
                self.add_address_details(_obj)
                self.add_dates_details(_obj)
                self.add_webs_details(_obj)
                self.add_fax_details(_obj)
                self.add_social_details(_obj)

                # try:
                #     business_card_mobile_qs = BusinessCardMobile.objects.filter(
                #         business_card_id=_obj.id,
                #     )
                # except ValueError:
                #     business_card_mobile_qs = None

                # try:
                #     for mobile_obj in business_card_mobile_qs:
                #         if mobile_obj.type == 'work':
                #             self.columns_list.append('mobile(Work)')
                #             self.row.append(str(mobile_obj.mobile_value))
                # except Exception as e:
                #     print(e)

            # Sheet header, first row
            row_num = 1
            # print(self.columns)
            # print(len(self.columns))

            # print("@@@@@@@@@@ ROW-QUERY-DATA @@@@@@@@@@@@@@@@@")
            # print(self.row_query_data)
            # print(len(self.row_query_data))

            # Assign the titles for each cell of the header
            for col_num, column_title in enumerate(self.columns_list, 1):
                cell = self.work_sheet.cell(row=row_num, column=col_num)
                cell.value = column_title
            # all query set values list data

            for _row_data in self.row_query_data:
                row_num += 1
                print("-------data-----")
                print(_row_data)
                # self.row = [

                #     # basic pay
                #     ############################
                #     str(getattr(_row_data, 'business_image', '0.0')),
                #     getattr(_row_data, 'first_name', '0.0'),

                #     "99090909090"

                # ]
                # print(row)
                # Assign the data for each cell of the row
                print(self.row)
                for col_num, cell_value in enumerate(self.row, 1):
                    cell = self.work_sheet.cell(row=row_num, column=col_num)
                    cell.value = cell_value

            # excel heading format
        excel_file_name = ''.join(['business_card_list', '.xls'])
        response = HttpResponse(
            content_type=self.sheet_type_content
        )
        response[
            'Content-Disposition'
        ] = 'attachment; filename="' + excel_file_name + '"'
        self.work_book_instance.save(response)
        return response
